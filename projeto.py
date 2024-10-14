import openpyxl
from time import sleep
from selenium import webdriver
# essa linha de importação acima permite abrir o navegador para automação
from selenium.webdriver.common.by import By
# essa linha serve para encontrar e interagir com elementos dentro de um pagina como inserir o cpf da pessoa

planilha_clientes = openpyxl.load_workbook('dados_clientes.xlsx')
# para carregar um workbook ou arquivo excell
pagina_clientes = planilha_clientes['Sheet1']
# pegar a indexação onde está a pagina dos clientes
driver = webdriver.Chrome()
# abrir o google chrome
driver.get('https://consultcpf-devaprender.netlify.app/')
# abrir o chrome e acessar o site que eu quero acessar
sleep(5)
try:
    arquivo_fechamento = 'planilha_fechamento.xlsx'
    planilha_fechamento = openpyxl.load_workbook(arquivo_fechamento)
    sleep(1)
    pagina_fechamento = planilha_fechamento['Sheet1']
    sleep(3)
except FileNotFoundError:
    print('arquivo "planilha_fechamento.xlsx" não encontrado')
except Exception as erro:
    print(f'erro motivo: {erro}')
    exit()
for linha in pagina_clientes.iter_rows(min_row=2, values_only=True):
    nome, valor, cpf, vencimento = linha
# a função iter_rows serve analisar linha por linha da pagina
# a função min_row= serve para dizer a partir de qual linha ele começa a analisar
# a função values_only=True serve para retornar apenas os valores que estão naquela linha
    # dar uma pausa para ter certeza que o site carregou
    try:
        campo_pesquisa = driver.find_element(By.XPATH, '//input[@id="cpfInput"]')
        sleep(1)
        # achar um elemento na página
        # xpath é a técnica para encontrar o elemento único na página, no caso é o cpf
        # método xpath //tag[@atributo='valor']
        campo_pesquisa.clear()
        # apaga os dados do cpf antes de digitar outro cpf
        campo_pesquisa.send_keys(cpf)
        sleep(1)
        # método que permite escrever dentro do campo determinado no elemento
        botao_consulta = driver.find_element(By.XPATH, '//button[@type="submit"]')
        sleep(1)
        botao_consulta.click()
        # metodo para dar um click no botão selecionado
        sleep(4)
        check_status = driver.find_element(By.XPATH, '//span[@id="statusLabel"]')
        sleep(1)
        if check_status.text == 'em dia':
            # variável.text extrai o texto do elemento achado
            data_pagamento = driver.find_element(By.XPATH, '//p[@id="paymentDate"]')
            sleep(1)
            metodo_pagamento = driver.find_element(By.XPATH, '//p[@id="paymentMethod"]')
            sleep(1)
            data_pagamentolimpo = data_pagamento.text.split()[3]
            metodo_pagamentolimpo = metodo_pagamento.text.split()[3]
            pagina_fechamento.append([nome, valor, cpf, vencimento, 'em dia', data_pagamentolimpo, metodo_pagamentolimpo])
        else:
            pagina_fechamento.append([nome, valor, cpf, vencimento, 'pendente'])
    except Exception as erro:
        print(f'erro na linha do cpf: {erro}')
    planilha_fechamento.save(arquivo_fechamento)
driver.quit()
print("Processo Concluído!")
